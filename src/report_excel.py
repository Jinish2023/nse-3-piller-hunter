"""
Excel tracker — single persistent sheet, appended every run, with
same-setup de-duplication.
"""
import os
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from . import config

_HEADER_STYLE_NAVY = "1F3864"
_GREEN_FILL = "C6E0B4"
_RED_FILL = "FF0000"
_BLUE_FILL = "9DC3E6"
_WHITE = "FFFFFF"

TRACKER_HEADERS = [
    "Serial No", "Date (IST)", "Stock", "Current Price", "RSI(14)",
    "Volume (today / 20d avg)", "Entry Price Range Low", "Entry Price Range High",
    "Stop Loss", "Stop Loss %", "Target1", "Target1 %", "Target2", "Target2 %",
    "Previous Close", "Entry? [Confirmed Breakout]", "Entry Price = Upper Bound",
    "Technical Details (3-Pillar Tags)", "Status", "Last Confirmed (IST)",
]


def _tracker_style_header(ws):
    header_font = Font(name="Arial", size=10, bold=True, color=_WHITE)
    header_fill = PatternFill("solid", fgColor=_HEADER_STYLE_NAVY)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, title in enumerate(TRACKER_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 40
    widths = {
        "A": 8, "B": 20, "C": 14, "D": 12, "E": 9, "F": 22,
        "G": 12, "H": 12, "I": 11, "J": 9, "K": 11, "L": 9, "M": 11, "N": 9, "O": 12,
        "P": 20, "Q": 16, "R": 55, "S": 12, "T": 20,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.sheet_view.showGridLines = False


def _tracker_load_or_create(path):
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        if config.TRACKER_SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(config.TRACKER_SHEET_NAME)
            _tracker_style_header(ws)
        return wb
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = config.TRACKER_SHEET_NAME
    _tracker_style_header(ws)
    return wb


def _tracker_existing_rows_by_symbol(ws):
    """Returns {symbol: [(row_idx, date_dt, entry_high, stop_loss), ...]} for fast lookup."""
    out = {}
    for row_idx in range(2, ws.max_row + 1):
        symbol = ws.cell(row=row_idx, column=3).value
        if not symbol:
            continue
        date_str = ws.cell(row=row_idx, column=2).value
        try:
            date_dt = datetime.strptime(date_str, "%d/%m/%Y %I:%M %p IST").replace(tzinfo=config.IST)
        except Exception:
            date_dt = None
        entry_high = ws.cell(row=row_idx, column=8).value
        stop_loss = ws.cell(row=row_idx, column=9).value
        out.setdefault(symbol, []).append((row_idx, date_dt, entry_high, stop_loss))
    return out


def _within_tolerance(a, b, pct):
    if a in (None, 0) or b in (None, 0):
        return False
    return abs(a - b) / abs(b) <= pct / 100.0


def append_candidates_to_tracker(candidates, run_dt=None, path=None):
    """
    De-duplication logic (best-practice for a swing-trade journal):
    - Same symbol logged within the last DEDUP_LOOKBACK_DAYS days, with entry-zone
      and stop-loss within DEDUP_TOLERANCE_PCT of the existing row -> treat as the
      SAME live setup. No new row is added; we just bump that row's
      "Last Confirmed (IST)" cell so you know it's still valid today.
    - Otherwise (first time seen, levels moved beyond tolerance, or it dropped off
      the scan for longer than the lookback window and reappeared) -> genuinely a
      new/changed setup, so it gets its own new row (Status = "New" or "Updated").
    This avoids the sheet bloating with near-identical rows every single day, while
    never silently hiding a real change in the setup.
    """
    path = path or config.TRACKER_XLSX_PATH
    run_dt = run_dt or datetime.now(config.IST)
    date_str = run_dt.strftime("%d/%m/%Y %I:%M %p IST")

    wb = _tracker_load_or_create(path)
    ws = wb[config.TRACKER_SHEET_NAME]
    existing = _tracker_existing_rows_by_symbol(ws)

    body_font = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    next_serial = ws.max_row - 1
    appended, refreshed = 0, 0

    for c in candidates:
        lv = c["levels"]
        symbol = c["symbol"]
        prev_close = c.get("prev_close", c["close"])
        entry_ok = lv.get("entry_confirmed", False)

        match_row = None
        for row_idx, date_dt, old_high, old_sl in existing.get(symbol, []):
            if date_dt and (run_dt - date_dt) <= timedelta(days=config.DEDUP_LOOKBACK_DAYS):
                if _within_tolerance(lv["entry_high"], old_high, config.DEDUP_TOLERANCE_PCT) and \
                   _within_tolerance(lv["stop_loss"], old_sl, config.DEDUP_TOLERANCE_PCT):
                    match_row = row_idx
                    break

        if match_row:
            ws.cell(row=match_row, column=20, value=date_str)
            refreshed += 1
            continue

        status = "Updated" if symbol in existing else "New"
        next_serial += 1
        r = next_serial + 1

        tags_text = "; ".join(c["tags"])
        vol_today_fmt = f"{c['volume_today']:.0f}"
        vol_avg_fmt = f"{c['volume_avg20']:.0f}"

        values = [
            next_serial, date_str, symbol, c["close"], round(c["rsi"]),
            f"{vol_today_fmt} / {vol_avg_fmt} ({c['vol_surge']:.1f}x)",
            lv["entry_low"], lv["entry_high"], lv["stop_loss"], lv["stop_loss_pct"] / 100.0,
            lv["target1"], lv["target1_pct"] / 100.0, lv["target2"], lv["target2_pct"] / 100.0,
            prev_close, "Yes" if entry_ok else "No", lv["entry_high"],
            tags_text, status, date_str,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = body_font
            cell.border = border
            cell.alignment = center_align

        for col_letter in ["D", "G", "H", "I", "K", "M", "O"]:
            ws[f"{col_letter}{r}"].number_format = "0.0"
        for col_letter in ["J", "L", "N"]:
            ws[f"{col_letter}{r}"].number_format = "+0.0%;-0.0%"

        ws[f"G{r}"].fill = PatternFill("solid", fgColor=_GREEN_FILL)
        ws[f"H{r}"].fill = PatternFill("solid", fgColor=_GREEN_FILL)
        ws[f"I{r}"].fill = PatternFill("solid", fgColor=_RED_FILL)
        ws[f"I{r}"].font = Font(name="Arial", size=10, color=_WHITE, bold=True)
        ws[f"J{r}"].fill = PatternFill("solid", fgColor=_RED_FILL)
        ws[f"J{r}"].font = Font(name="Arial", size=10, color=_WHITE, bold=True)
        for col_letter in ["K", "L", "M", "N"]:
            ws[f"{col_letter}{r}"].fill = PatternFill("solid", fgColor=_BLUE_FILL)
        ws[f"O{r}"].fill = PatternFill("solid", fgColor=_GREEN_FILL)
        ws[f"R{r}"].alignment = wrap_align

        appended += 1

    wb.save(path)
    print(f"  Tracker updated: {appended} new row(s), {refreshed} existing row(s) refreshed -> {path}")
    return appended, refreshed
